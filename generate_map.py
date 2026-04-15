#!/usr/bin/env python3
"""
generate_map.py — Genera card_estudiar_fora.svg amb projecció azimutal
centrada a Barcelona (41.39°N, 2.17°E), estirada a el·lipse 16:9.

Escala radial potència (POWER < 1) per expandir Europa i comprimir Oceania.

Ús: python3 generate_map.py
Entrades:
  - universitats vs centres UPF.xlsx  (font de veritat de la llista d'Enginyeria,
    amb una "x" a la columna "Escola d'Enginyeria")
  - destinacions_upf.csv  (per a coordenades, ciutats i continents — join per Codi)
  - ne_110m_land.geojson  (auto-descarregat)
Sortida: card_estudiar_fora.svg

Dependència: openpyxl (per a llegir el fitxer xlsx).
"""

import math
import csv
import json
import os
import re
import urllib.request

try:
    from openpyxl import load_workbook
except ImportError:
    raise SystemExit("Cal openpyxl: pip install openpyxl")

# ─── Constants ────────────────────────────────────────────────────────────────

VIEWBOX_W, VIEWBOX_H = 800, 420
CENTER_X, CENTER_Y = 388, 210

# Projection center: Barcelona
CENTER_LAT, CENTER_LON = 41.3874, 2.1686

# Elliptical globe: fills the 16:9 viewBox with small margins
# ry constrained by height, rx = ry * (W/H) to match aspect ratio
GLOBE_RY = 195                          # vertical semi-axis (15px margin)
GLOBE_RX = GLOBE_RY * VIEWBOX_W / VIEWBOX_H  # ≈ 371 horizontal semi-axis

# Power scaling: p<1 expands near distances (Europe), compresses far (Oceania)
# p=1.0 → equidistant, p=0.6 → moderate fisheye, p=0.5 → strong fisheye
POWER = 0.5

# Scale: full globe (π^POWER radians) maps to each semi-axis
SCALE_X = GLOBE_RX / (math.pi ** POWER)
SCALE_Y = GLOBE_RY / (math.pi ** POWER)
EARTH_R = 6371  # km

# Colors (matching current SVG)
BG_DARK = "#0b1526"
BG_LIGHT = "#142236"
MARKER_COLOR = "#e8475f"
BCN_COLOR = "#c8102e"

# Distance rings in km
DISTANCE_RINGS = [2000, 5000, 10000, 15000]

# File paths
CSV_PATH = "destinacions_upf.csv"
XLSX_PATH = "/Users/alfonsomartinez/Library/CloudStorage/Dropbox/Downloads/universitats vs centres UPF.xlsx"
ENG_COLUMN = "Escola d'Enginyeria"
GEOJSON_URL = "https://raw.githubusercontent.com/nvkelso/natural-earth-vector/master/geojson/ne_110m_land.geojson"
GEOJSON_CACHE = "ne_110m_land.geojson"
OUTPUT = "card_estudiar_fora.svg"


def normalize_codi(s):
    """Normalize Erasmus code: collapse all whitespace (incl. nbsp) to single space, uppercase."""
    if not s:
        return ""
    return re.sub(r"\s+", " ", str(s).replace("\xa0", " ")).strip().upper()

# Cities to label (lowercase) — curated for geographic spread in azimuthal view
# European cities (< 20px from center) are unlabeled to avoid overlap with BARCELONA
LABEL_CITIES = {
    # Oceania (farthest — show global reach)
    "sydney", "melbourne", "brisbane",
    # East Asia
    "tokyo", "seoul",
    # South Asia
    "bangalore",
    # South America
    "buenos aires", "santiago de xile", "rio de janeiro",
    # North America
    "vancouver", "chicago", "ciudad de méxico",
    # Medium-distance Europe (> 20px from center with POWER=0.6)
    "reykjavik", "moscow",
    # Close Europe (iconic — user requested)
    "madrid", "münchen",
    "lisboa", "roma", "milano",
}

# Normalize city names for deduplication (handles València/Valencia etc.)
CITY_NORMALIZE = {
    "valència": "valencia",
}

# SICUE (Spanish national mobility) destinations — not in the international
# offer xlsx. Hard-coded here. Each entry: name → (city, country, lat, lon, continent).
SICUE_DESTINATIONS = {
    "UNIVERSIDAD CARLOS III DE MADRID":           ("Getafe",     "Espanya", 40.3170, -3.7620, "Europa"),
    "UNIVERSIDAD DE A CORUÑA":                    ("A Coruña",   "Espanya", 43.3300, -8.4130, "Europa"),
    "UNIVERSIDAD DE OVIEDO":                      ("Oviedo",     "Espanya", 43.3590, -5.8520, "Europa"),
    "UNIVERSIDAD DE VALENCIA (ESTUDI GENERAL)":   ("València",   "Espanya", 39.5100, -0.4210, "Europa"),
    "UNIVERSIDAD POLITÉCNICA DE VALENCIA":        ("València",   "Espanya", 39.4820, -0.3470, "Europa"),
    "UNIVERSITAT DE LES ILLES BALEARS":           ("Palma",      "Espanya", 39.6380,  2.6500, "Europa"),
}

# Manual coordinates for universities whose Codi is a country-level placeholder
# (USA999, CAN999, JPN999, KOR999, THA999, TWN999, AUS999) and therefore not
# in destinacions_upf.csv. Keyed by uppercase university name.
EXTRA_DESTINATIONS = {
    "THE UNIVERSITY OF SYDNEY":     ("Sydney",       "Austràlia",                -33.8888, 151.1872, "Oceania"),
    "DALHOUSIE UNIVERSITY":         ("Halifax",      "Canadà",                    44.6366, -63.5917, "Amèrica del Nord"),
    "SIMON FRASER UNIVERSITY":      ("Burnaby",      "Canadà",                    49.2768, -122.9180, "Amèrica del Nord"),
    "KOREA UNIVERSITY":             ("Seoul",        "Corea del Sud",             37.5894, 127.0327, "Àsia"),
    "BARNARD COLLEGE":              ("Nova York",    "Estats Units d'Amèrica",    40.8093, -73.9617, "Amèrica del Nord"),
    "DREXEL UNIVERSITY":            ("Filadèlfia",   "Estats Units d'Amèrica",    39.9555, -75.1882, "Amèrica del Nord"),
    "REED COLLEGE":                 ("Portland",     "Estats Units d'Amèrica",    45.4810, -122.6294, "Amèrica del Nord"),
    "UNIVERSITY OF RICHMOND":       ("Richmond",     "Estats Units d'Amèrica",    37.5754, -77.5410, "Amèrica del Nord"),
    "WEST VIRGINIA UNIVERSITY":     ("Morgantown",   "Estats Units d'Amèrica",    39.6358, -79.9555, "Amèrica del Nord"),
    "KEIO UNIVERSITY":              ("Tokyo",        "Japó",                      35.6537, 139.7434, "Àsia"),
    "CHULALONGKORN UNIVERSITY":     ("Bangkok",      "Tailàndia",                 13.7383, 100.5328, "Àsia"),
    "NATIONAL TAIWAN UNIVERSITY":   ("Taipei City",  "Taiwan",                    25.0173, 121.5398, "Àsia"),
}

# ─── Projection (precomputed center trig) ────────────────────────────────────

_phi0 = math.radians(CENTER_LAT)
_lam0 = math.radians(CENTER_LON)
_sp0 = math.sin(_phi0)
_cp0 = math.cos(_phi0)


def project(lat_deg, lon_deg):
    """Azimuthal projection with power scaling → (svg_x, svg_y).
    k = c^POWER / sin(c) instead of c / sin(c) for non-linear radial scale."""
    phi = math.radians(lat_deg)
    lam = math.radians(lon_deg)
    dl = lam - _lam0
    sp, cp, cdl = math.sin(phi), math.cos(phi), math.cos(dl)
    cos_c = _sp0 * sp + _cp0 * cp * cdl
    cos_c = max(-1.0, min(1.0, cos_c))
    c = math.acos(cos_c)
    if c < 1e-10:
        return CENTER_X, CENTER_Y
    k = (c ** POWER) / math.sin(c)
    x_raw = k * cp * math.sin(dl)
    y_raw = -(k * (_cp0 * sp - _sp0 * cp * cdl))
    return CENTER_X + x_raw * SCALE_X, CENTER_Y + y_raw * SCALE_Y


def ang_dist(lat_deg, lon_deg):
    """Angular distance from center in radians."""
    phi = math.radians(lat_deg)
    dl = math.radians(lon_deg) - _lam0
    cos_c = _sp0 * math.sin(phi) + _cp0 * math.cos(phi) * math.cos(dl)
    return math.acos(max(-1.0, min(1.0, cos_c)))


def km_to_radii(km):
    """Convert distance in km to elliptical SVG radii (rx, ry) with power scaling."""
    rad = km / EARTH_R
    scaled = rad ** POWER
    return scaled * SCALE_X, scaled * SCALE_Y


# ─── Data loading ────────────────────────────────────────────────────────────

def load_engineering_rows():
    """Read the xlsx and return [(codi, university_name, country), ...] for
    every row marked as offered to Engineering."""
    wb = load_workbook(XLSX_PATH, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    code_idx = header.index("Codi universitat")
    eng_idx = header.index(ENG_COLUMN)
    country_idx = header.index("País")
    out = []
    for r in rows[1:]:
        if not r or not r[0]:
            continue
        eng = r[eng_idx]
        if eng and str(eng).strip().lower() in ("x", "s", "sí", "si", "yes", "1", "true"):
            out.append((normalize_codi(r[code_idx]), str(r[0]).strip(), str(r[country_idx] or "").strip()))
    return out


def load_csv_lookup():
    """Index destinacions_upf.csv by normalized Codi → coords + city/country/continent."""
    by_codi = {}
    with open(CSV_PATH, encoding="utf-8") as f:
        reader = csv.reader(f)
        next(reader)
        for row in reader:
            if len(row) < 8:
                continue
            lat_s, lon_s = row[2].strip(), row[3].strip()
            if not lat_s or not lon_s:
                continue
            try:
                lat, lon = float(lat_s), float(lon_s)
            except ValueError:
                continue
            codi = normalize_codi(row[0])
            if not codi:
                continue
            by_codi[codi] = {
                "city": row[5].strip(),
                "country": row[6].strip(),
                "continent": row[7].strip(),
                "lat": lat, "lon": lon,
            }
    return by_codi


def load_destinations():
    """Build the Engineering destination list from the xlsx, joining coords
    from destinacions_upf.csv and falling back to EXTRA_DESTINATIONS for
    placeholder codes (USA999, CAN999, etc.).

    Returns (full_list, dedup_for_map):
      - full_list: every offered destination (used for legend counts)
      - dedup_for_map: deduplicated by city, used for plotting markers + lines
    """
    eng_rows = load_engineering_rows()
    csv_by_codi = load_csv_lookup()

    full = []
    unmatched = []
    for codi, uni_name, _ in eng_rows:
        if codi in csv_by_codi:
            d = dict(csv_by_codi[codi]); d["uni"] = uni_name
            full.append(d); continue
        key = uni_name.upper()
        if key in EXTRA_DESTINATIONS:
            city, c_country, lat, lon, continent = EXTRA_DESTINATIONS[key]
            full.append({"uni": uni_name, "city": city, "country": c_country,
                         "continent": continent, "lat": lat, "lon": lon})
            continue
        unmatched.append((codi, uni_name))

    if unmatched:
        print(f"⚠ Universitats sense coordenades ({len(unmatched)}):")
        for codi, name in unmatched:
            print(f"    [{codi}] {name}")

    # Add SICUE (Spanish national mobility) destinations
    for uni_name, (city, country, lat, lon, continent) in SICUE_DESTINATIONS.items():
        full.append({"uni": uni_name, "city": city, "country": country,
                     "continent": continent, "lat": lat, "lon": lon})
    print(f"Afegides {len(SICUE_DESTINATIONS)} destinacions SICUE")

    # Dedup by city for plotting
    dests = []
    seen_cities = set()
    for d in full:
        ck = CITY_NORMALIZE.get(d["city"].lower(), d["city"].lower())
        if ck in seen_cities:
            continue
        seen_cities.add(ck)
        dests.append(d)

    return full, dests


def load_land():
    """Download (if needed) and parse Natural Earth 110m land GeoJSON."""
    if not os.path.exists(GEOJSON_CACHE):
        print("Descarregant Natural Earth 110m land...")
        urllib.request.urlretrieve(GEOJSON_URL, GEOJSON_CACHE)
        print(f"  Desat a {GEOJSON_CACHE}")

    with open(GEOJSON_CACHE, encoding="utf-8") as f:
        data = json.load(f)

    rings = []
    for feat in data["features"]:
        g = feat["geometry"]
        if g["type"] == "Polygon":
            rings.extend(g["coordinates"])
        elif g["type"] == "MultiPolygon":
            for poly in g["coordinates"]:
                rings.extend(poly)
    return rings


# ─── Project coastline rings with antipode handling ──────────────────────────

def project_ring(ring, max_ang=2.97):
    """Project a GeoJSON ring → list of segments [(x,y), ...].
    Splits at points near antipode (>170°) or with big SVG-space jumps."""
    segments = []
    seg = []

    for lon, lat in ring:
        if ang_dist(lat, lon) > max_ang:
            if len(seg) >= 2:
                segments.append(seg)
            seg = []
            continue

        x, y = project(lat, lon)

        # Check for wild jumps (antipodal wraparound)
        # Threshold scaled for elliptical projection (~250px)
        if seg:
            dx, dy = x - seg[-1][0], y - seg[-1][1]
            if dx * dx + dy * dy > 62500:  # > 250px
                if len(seg) >= 2:
                    segments.append(seg)
                seg = []

        seg.append((x, y))

    if len(seg) >= 2:
        segments.append(seg)
    return segments


# ─── SVG Generation ──────────────────────────────────────────────────────────

def build_svg(full, dests, land_rings, interactive=False):
    """Generate the complete SVG string. full = all destinations (legend
    counts), dests = deduped-by-city subset (plotted markers + lines).
    If interactive=True, each marker gets a <title> child listing the
    universities at that city (browsers render as a hover tooltip)."""
    # Index unis by normalized city name → list of university display names
    from collections import defaultdict
    unis_by_city = defaultdict(list)
    for d in full:
        ck = CITY_NORMALIZE.get(d["city"].lower(), d["city"].lower())
        if "uni" in d:
            unis_by_city[ck].append(d["uni"])
        else:
            # Fallback: anonymous "Universitat de X" placeholder when we
            # joined via CSV and don't have a name in the entry
            unis_by_city[ck].append("Universitat")

    s = []
    a = s.append

    # ── Header & defs ──
    a('<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 800 420"'
      ' preserveAspectRatio="xMidYMid slice">')
    a("<defs>")
    a(f'  <linearGradient id="bg" x1="0" y1="0" x2="1" y2="1">'
      f'<stop offset="0%" stop-color="{BG_DARK}"/>'
      f'<stop offset="100%" stop-color="{BG_LIGHT}"/></linearGradient>')
    a('  <filter id="gl"><feGaussianBlur stdDeviation="1.5" result="b"/>'
      "<feMerge><feMergeNode in=\"b\"/>"
      '<feMergeNode in="SourceGraphic"/></feMerge></filter>')
    a(f'  <radialGradient id="pu" cx="50%" cy="50%" r="50%">'
      f'<stop offset="0%" stop-color="{BCN_COLOR}" stop-opacity="0.7"/>'
      f'<stop offset="100%" stop-color="{BCN_COLOR}" stop-opacity="0"/>'
      f"</radialGradient>")
    a(f'  <clipPath id="gc"><ellipse cx="{CENTER_X}" cy="{CENTER_Y}"'
      f' rx="{GLOBE_RX:.1f}" ry="{GLOBE_RY}"/></clipPath>')
    a("</defs>")

    # ── Background ──
    a(f'<rect width="{VIEWBOX_W}" height="{VIEWBOX_H}" fill="url(#bg)"/>')

    # ── Globe boundary ellipse ──
    a(f'<ellipse cx="{CENTER_X}" cy="{CENTER_Y}"'
      f' rx="{GLOBE_RX:.1f}" ry="{GLOBE_RY}"'
      f' fill="none" stroke="#fff" stroke-width="0.5" opacity="0.08"/>')

    # ── Distance rings (ellipses) ──
    a('<g opacity="0.05" stroke="#fff" stroke-width="0.3" fill="none">')
    for km in DISTANCE_RINGS:
        rx, ry = km_to_radii(km)
        a(f'  <ellipse cx="{CENTER_X}" cy="{CENTER_Y}"'
          f' rx="{rx:.1f}" ry="{ry:.1f}"/>')
    a("</g>")

    # ── Distance ring labels (top of each ring) ──
    a('<g fill="#fff" font-family="Arial" font-size="4"'
      ' opacity="0.15" text-anchor="middle">')
    for km in DISTANCE_RINGS:
        _, ry = km_to_radii(km)
        label = f"{km // 1000}.000 km"
        a(f'  <text x="{CENTER_X}" y="{CENTER_Y - ry - 2:.1f}">{label}</text>')
    a("</g>")

    # ── Continent silhouettes (clipped to globe) ──
    a('<g clip-path="url(#gc)">')
    for ring in land_rings:
        for seg in project_ring(ring):
            pts = " ".join(f"{x:.1f},{y:.1f}" for x, y in seg)
            if len(seg) >= 3:
                a(f'  <polygon points="{pts}" fill="#fff" opacity="0.08"'
                  f' stroke="#fff" stroke-width="0.3" stroke-opacity="0.1"/>')
            else:
                a(f'  <polyline points="{pts}" fill="none"'
                  f' stroke="#fff" stroke-width="0.3" opacity="0.1"/>')
    a("</g>")

    # ── Lines from Barcelona (straight in azimuthal equidistant!) ──
    a(f'<g stroke="{BCN_COLOR}" stroke-width="0.35" opacity="0.15">')
    for d in dests:
        x, y = project(d["lat"], d["lon"])
        a(f'  <line x1="{CENTER_X}" y1="{CENTER_Y}"'
          f' x2="{x:.1f}" y2="{y:.1f}"/>')
    a("</g>")

    # ── Destination markers with glow ──
    a('<g filter="url(#gl)">')
    for d in dests:
        x, y = project(d["lat"], d["lon"])
        if interactive:
            ck = CITY_NORMALIZE.get(d["city"].lower(), d["city"].lower())
            unis = unis_by_city.get(ck, [])
            title = f'{d["city"]} — {", ".join(unis)}' if unis else d["city"]
            title = title.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
            a(f'  <circle cx="{x:.1f}" cy="{y:.1f}" r="2.5"'
              f' fill="{MARKER_COLOR}" opacity="0.9">')
            a(f'    <title>{title}</title>')
            a('  </circle>')
        else:
            a(f'  <circle cx="{x:.1f}" cy="{y:.1f}" r="2.5"'
              f' fill="{MARKER_COLOR}" opacity="0.9"/>')
    a("</g>")

    # ── Barcelona center with pulse animation ──
    a(f'<circle cx="{CENTER_X}" cy="{CENTER_Y}" r="4.5"'
      f' fill="{BCN_COLOR}" filter="url(#gl)"/>')
    a(f'<circle cx="{CENTER_X}" cy="{CENTER_Y}" r="11" fill="url(#pu)">')
    a('  <animate attributeName="r" values="6;13;6"'
      ' dur="2.5s" repeatCount="indefinite"/>')
    a('  <animate attributeName="opacity" values="0.6;0.1;0.6"'
      ' dur="2.5s" repeatCount="indefinite"/>')
    a("</circle>")
    a(f'<text x="{CENTER_X}" y="{CENTER_Y - 9}" text-anchor="middle"'
      f' fill="#fff" font-size="7.5" font-family="Arial"'
      f' font-weight="bold" opacity="0.95">BARCELONA</text>')

    # ── City labels ──
    a('<g fill="#fff" font-family="Arial" font-size="5.5" opacity="0.5">')
    for d in dests:
        if d["city"].lower() not in LABEL_CITIES:
            continue
        x, y = project(d["lat"], d["lon"])
        # text-anchor based on position relative to center
        if x > CENTER_X:
            anchor, lx = "start", x + 5
        else:
            anchor, lx = "end", x - 5
        a(f'  <text x="{lx:.1f}" y="{y - 4:.1f}"'
          f' text-anchor="{anchor}">{d["city"]}</text>')
    a("</g>")

    # ── No legend baked into the SVG (the HTML hero overlay carries the
    # numbers, which Liferay can render in CA/ES/EN per page locale). ──

    a("</svg>")
    return "\n".join(s)


# ─── Main ────────────────────────────────────────────────────────────────────

def main():
    print("=== Generador mapa azimutal (el·líptic 16:9) ===\n")
    print(f"El·lipse: rx={GLOBE_RX:.1f} ry={GLOBE_RY}")
    print(f"Escala potència: p={POWER} (1.0=equidistant, <1=fisheye)\n")

    # Load destinations: full = xlsx truth (59), dests = deduped for plotting
    full, dests = load_destinations()
    print(f"Destinacions Enginyeria (xlsx): {len(full)}  |  punts al mapa (dedup ciutat): {len(dests)}")
    for d in sorted(dests, key=lambda x: x["city"]):
        x, y = project(d["lat"], d["lon"])
        print(f"  {d['city']:25s} {d['country']:20s}"
              f" → ({x:6.1f}, {y:5.1f})")

    countries = set(d["country"] for d in full)
    print(f"\nPaïsos (xlsx): {len(countries)}")

    # Load coastlines
    land = load_land()
    print(f"Polígons costa: {len(land)}")

    # Generate both SVGs: static (default) and interactive (with hover tooltips)
    svg = build_svg(full, dests, land, interactive=False)
    with open(OUTPUT, "w", encoding="utf-8") as f:
        f.write(svg)
    print(f"\n✓ SVG estàtic generat: {OUTPUT} ({len(svg):,} bytes)")

    interactive_path = OUTPUT.replace(".svg", "_interactive.svg")
    svg_i = build_svg(full, dests, land, interactive=True)
    with open(interactive_path, "w", encoding="utf-8") as f:
        f.write(svg_i)
    print(f"✓ SVG interactiu generat: {interactive_path} ({len(svg_i):,} bytes)")

    # Show which labels were matched
    labeled = [d["city"] for d in dests if d["city"].lower() in LABEL_CITIES]
    print(f"Etiquetes mostrades: {len(labeled)}")
    for name in sorted(labeled):
        print(f"  {name}")


if __name__ == "__main__":
    main()
